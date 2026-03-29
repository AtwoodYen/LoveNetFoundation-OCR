"""
從合併後的 OCR 結果產生 Excel（含手寫區塊分頁）。
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


# 模型若單獨標記手寫區塊時會使用的 label（印刷文字多為 text，不併入此表）
_HANDWRITING_LABELS = frozenset(
    {
        "handwriting",
        "handwritten",
        "hand_writing",
    }
)


def _plain_text(block_content: str) -> str:
    if not block_content:
        return ""
    # 粗略移除圖片區塊的 HTML
    s = re.sub(r"<[^>]+>", " ", block_content)
    return " ".join(s.split()).strip()


def build_task_excel(result_data: Dict[str, Any]) -> bytes:
    """
    依 merged.json 結構（full_markdown, layout）建立 xlsx。
    layout 項目應含 layout_type, block_content, page_index, block_id。
    """
    layout: List[Dict[str, Any]] = result_data.get("layout") or []
    meta = result_data.get("metadata") or {}

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "全部區塊"
    headers = ["序號", "頁碼", "區塊類型", "文字內容（純文字）"]
    ws_all.append(headers)
    for c in range(1, len(headers) + 1):
        ws_all.cell(1, c).font = Font(bold=True)

    for i, block in enumerate(layout, start=1):
        page = block.get("page_index", "")
        btype = block.get("layout_type", "")
        raw = block.get("block_content", "") or ""
        ws_all.append([i, page, btype, _plain_text(raw)])

    ws_hw = wb.create_sheet("手寫文字")
    ws_hw.append(["序號", "頁碼", "區塊類型", "文字內容"])
    for c in range(1, 5):
        ws_hw.cell(1, c).font = Font(bold=True)

    hw_seq = 0
    for block in layout:
        btype = (block.get("layout_type") or "").lower()
        if btype not in _HANDWRITING_LABELS:
            continue
        hw_seq += 1
        page = block.get("page_index", "")
        raw = block.get("block_content", "") or ""
        ws_hw.append([hw_seq, page, btype, _plain_text(raw)])

    if hw_seq == 0:
        ws_hw.append(
            [
                "",
                "",
                "",
                "尚無 label 為 handwriting/handwritten 的區塊；若模型將手寫併入 text，請於「全部區塊」篩選 type=text。",
            ]
        )
        ws_hw["D2"].alignment = Alignment(wrap_text=True)

    # 摘要
    ws_meta = wb.create_sheet("摘要")
    ws_meta.append(["項目", "值"])
    ws_meta.cell(1, 1).font = Font(bold=True)
    ws_meta.cell(1, 2).font = Font(bold=True)
    ws_meta.append(["原始檔名", meta.get("original_filename", "")])
    ws_meta.append(["總頁數", meta.get("total_pages", "")])
    ws_meta.append(["區塊數", len(layout)])

    od = result_data.get("offering_display")
    if isinstance(od, dict):
        for row in od.get("fields") or []:
            if isinstance(row, dict) and row.get("label") and row.get("value"):
                ws_meta.append([row.get("label", ""), row.get("value", "")])
        for item in od.get("checked_items") or []:
            if item:
                ws_meta.append(["勾選項目", str(item)])

    for row in ws_all.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws_hw.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_all.column_dimensions["A"].width = 8
    ws_all.column_dimensions["B"].width = 8
    ws_all.column_dimensions["C"].width = 14
    ws_all.column_dimensions["D"].width = 80
    ws_hw.column_dimensions["D"].width = 80

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
